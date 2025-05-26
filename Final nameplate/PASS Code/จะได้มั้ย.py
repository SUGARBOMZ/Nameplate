# --------------------  extract-excel.py  (FULL FILE – 27 Apr 2025)  --------------------
import os, base64, json, re, io, imghdr, requests, pandas as pd, streamlit as st
API_KEY = "AIzaSyDb8iBV1EWqLvjheG_44gh3vQHfpmYGOCI"
from PIL import Image
from openpyxl import load_workbook

def fill_template_from_validated(validated_path, template_path,
                                 key_col_template='ASSETNUM',
                                 key_col_validated='Correct_POWTR_CODE',
                                 sheet_name='AssetAttr') -> io.BytesIO:
    # 1) อ่าน df_validate
    df = pd.read_excel(validated_path)
    df = df[df['Is_Correct']==True].set_index(key_col_validated)

    # 2) โหลด workbook + sheet
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # 3) ลบแถวแรก (group header)
    ws.delete_rows(1)

    # 4) อ่าน header map: name→col_index
    header = next(ws.iter_rows(min_row=1, max_row=1))
    cols = {cell.value: cell.column for cell in header if cell.value}

    # 5) วนแต่ละแถว (จาก row 2) เติมข้อมูล
    for row in ws.iter_rows(min_row=2):
        asset = row[cols['ASSETNUM']-1].value
        if asset in df.index:
            rec = df.loc[asset]
            # ASSETNUM เป็น key ใหม่ (Correct_POWTR_CODE)
            row[cols['ASSETNUM']-1].value = rec.name
            # SITEID ← Plant, HIERARCHYPATH ← Location Description
            row[cols['SITEID']-1].value = rec['Plant']
            row[cols['HIERARCHYPATH']-1].value = rec['Location Description']

    # 6) เซฟเป็น BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# --------------------------------------------------------------------------- #
# 1)  Utilities                                                               #
# --------------------------------------------------------------------------- #
def encode_image(file) -> tuple[str, str]:
    """Convert an uploaded image file to (base64‑string, mime‑type) for Gemini."""
    raw = file.getvalue()
    kind = imghdr.what(None, raw) or 'jpeg'
    mime = f"image/{kind}"
    return base64.b64encode(raw).decode('utf-8'), mime


def _kv_from_text(txt: str) -> float | None:
    """
    Return the highest **system voltage** (kV) found in *txt*.

    • Accept “… 525000 V”, “34.5 kV”, “220 kV”, etc.  
    • Ignore values that are clearly power/current ratings (kVA, A, kA, VA).  
    • Skip numbers near “BIL” or “IMPULSE”.  
    • Discard absurdly large values > 1500 kV.
    """
    txt_u = txt.upper()
    best = None

    # split on '/', ',', ';' to isolate “33/30000/309000 kV” cases
    for chunk in re.split(r'[\/,;]', txt_u):
        chunk = chunk.strip()

        # skip chunks that mention KVA / VA / KA / A
        if re.search(r'\bK?VA\b|\bKA\b|\bAMP|\bA\b', chunk):
            continue

        # skip if near BIL / IMPULSE
        if 'BIL' in chunk or 'IMPULSE' in chunk:
            continue

        for m in re.finditer(r'(\d+(?:\.\d+)?)\s*([K]?V)(?![A-Z])', chunk):
            n = float(m.group(1))
            unit = m.group(2).upper()
            kv = n if unit == 'KV' else n / 1000
            if kv > 1500:          # absurdly high → ignore
                continue
            best = kv if best is None else max(best, kv)

    return best


# --------------------------------------------------------------------------- #
# 2)  Prompt generator                    #
# --------------------------------------------------------------------------- #
def generate_prompt_from_excel(excel_file):
    """
    Read an Excel list of attributes + (optionally) units, then build a Thai prompt
    telling Gemini to extract those exact fields in JSON.
    """


    # ----- read Excel whether it has a header row or not -----
    try:
        df = pd.read_excel(excel_file)
        first_col = df.columns[0]
        is_numeric_header = isinstance(first_col, (int, float))
        if is_numeric_header:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, header=None)
            df.columns = ['attribute_name'] + [f'col_{i}' for i in range(1, len(df.columns))]
            st.info("ตรวจพบไฟล์ไม่มีหัวคอลัมน์ – กำลังปรับให้อ่านได้")
    except Exception as e:
        excel_file.seek(0)
        df = pd.read_excel(excel_file, header=None)
        df.columns = ['attribute_name'] + [f'col_{i}' for i in range(1, len(df.columns))]
        st.warning(f"อ่านไฟล์แบบมีหัวคอลัมน์ไม่ได้: {e}  → ใช้โหมดไม่มีหัว")

    st.write("คอลัมน์ที่พบ:", list(df.columns))

    attribute_col = 'attribute_name'
    if attribute_col not in df.columns:
        for c in ['attribute_name', 'attribute', 'name', 'attributes',
                  'Attribute', 'ATTRIBUTE', 'field', 'Field', 'FIELD']:
            if c in df.columns:
                attribute_col = c; break
        if attribute_col not in df.columns:
            attribute_col = df.columns[0]
            st.warning(f"ไม่พบคอลัมน์ชื่อ attribute ที่รู้จัก – ใช้คอลัมน์ '{attribute_col}' แทน")

    unit_col = None
    for c in ['unit_of_measure', 'unit', 'Unit', 'UNIT', 'uom', 'UOM',
              'unit of measure', 'Unit of Measure']:
        if c in df.columns:
            unit_col = c; break

    if unit_col is None and len(df.columns) > 1:
        potential = df.columns[1]
        sample = df[potential].dropna().astype(str).tolist()[:10]
        if any(any(k in v for k in ['kg', 'V', 'A', 'kV', 'kVA', 'C', '°C',
                                    'mm', 'cm', 'm', '%']) for v in sample):
            unit_col = potential
            st.info(f"ตรวจพบคอลัมน์ '{potential}' อาจเป็นหน่วยวัด")

    prompt_parts = ["""กรุณาสกัดข้อมูลทั้งหมดจากรูปภาพนี้และแสดงผลในรูปแบบ JSON ที่มีโครงสร้างชัดเจน โดยใช้ key เป็นภาษาอังกฤษและ value เป็นข้อมูลที่พบ
ให้ return ค่า attributes กลับด้วยค่า attribute เท่านั้นห้าม return เป็น index เด็ดขาดและไม่ต้องเอาค่า index มาด้วย ให้ระวังเรื่อง voltage high side หน่วยต้องเป็น V หรือ kV เท่านั้น
โดยเอาเฉพาะ attributes ดังต่อไปนี้\n"""]

    for i, row in df.iterrows():
        attr = str(row[attribute_col]).strip()
        if pd.isna(attr) or attr == '':
            continue
        if unit_col and unit_col in df.columns and pd.notna(row[unit_col]) and str(row[unit_col]).strip():
            prompt_parts.append(f"{i+1}: {attr} [{row[unit_col]}]")
        else:
            prompt_parts.append(f"{i+1}: {attr}")

    prompt_parts.append("\nหากไม่พบข้อมูลสำหรับ attribute ใด ให้ใส่ค่า - แทน ไม่ต้องเดาค่า และให้รวม attribute และหน่วยวัดไว้ในค่าที่ส่งกลับด้วย")
    return "\n".join(prompt_parts)



# --------------------------------------------------------------------------- #
# 3)  Gemini API call                                                         #
# --------------------------------------------------------------------------- #
def extract_data_from_image(api_key: str, img_b64: str, mime: str, prompt: str) -> str:
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-04-17:generateContent?key={api_key}"
    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inlineData": {"mimeType": mime, "data": img_b64}}
            ]
        }],
        "generationConfig": {"temperature": 0.2, "topP": 0.85, "maxOutputTokens": 9000}
    }
    r = requests.post(url, headers={"Content-Type": "application/json"}, data=json.dumps(payload))
    if r.ok and r.json().get('candidates'):
        return r.json()['candidates'][0]['content']['parts'][0]['text']
    return f"API ERROR {r.status_code}: {r.text}"


# --------------------------------------------------------------------------- #
# 4)  POWTR‑CODE generator                                                    #
# --------------------------------------------------------------------------- #
def generate_powtr_code(extracted: dict) -> str:
    try:
# 1) Phase
        phase = '3'
        if any(any(k in str(v).upper() for k in ('1PH', '1-PH', 'SINGLE'))
               for v in extracted.values()):
            phase = '1'

        # 2) Voltage level
        high_kv = None
        for k, v in extracted.items():
            if any(t in k.upper() for t in ('VOLT', 'HV', 'LV', 'RATED', 'SYSTEM')):
                kv = _kv_from_text(str(v))
                if kv is not None:
                    high_kv = kv if high_kv is None else max(high_kv, kv)

        if high_kv is None:
            v_char = '-'
        elif high_kv > 765:
            return 'POWTR-3-OO'
        elif high_kv >= 345:
            v_char = 'E'
        elif high_kv >= 100:
            v_char = 'H'
        elif high_kv >= 1:
            v_char = 'M'
        else:
            v_char = 'L'

        # 3) Type → default = '-'  (เมื่อตรวจไม่เจอทั้ง DRY และ OIL)
        t_char = '-'
        for v in extracted.values():
            u = str(v).upper()
            if 'DRY' in u:
                t_char = 'D'
                break
            # ตรวจหา oil-based cooling class (OIL, ONAN, OFAF, ...)
            if any(kw in u for kw in ('OIL', 'ONAN', 'OFAF', 'OA', 'FOA')):
                t_char = 'O'
                break

        # 4) Tap‑changer (เหมือนเดิม)
        tap_char = 'F'
        for v in extracted.values():
            u = str(v).upper()
            if any(x in u for x in ('ON‑LOAD', 'ON-LOAD', 'OLTC')):
                tap_char = 'O'
                break
            if any(x in u for x in ('OFF‑LOAD', 'OFF-LOAD', 'FLTC', 'OCTC')):
                tap_char = 'F'

        code = f'POWTR-{phase}{v_char}{t_char}{tap_char}'
        prefix = code.split('-',1)[0]   # จะได้ 'POWTR'
        return f"{prefix} \\ {code}"
    except Exception:
        return 'ไม่สามารถระบุได้'

def add_powtr_codes(results):
    for r in results:
        d = r.get('extracted_data', {})
        if isinstance(d, dict) and not any(k in d for k in ('error','raw_text')):
            # ใส่ POWTR_CODE เข้าไป
            code = generate_powtr_code(d)
            d['POWTR_CODE'] = code
    return results

def split_value_unit(val: object) -> tuple[str, str]:
    """
    แยก '123 kg' -> ('123','kg')
    ถ้าไม่มีหน่วย ให้คืน (str(val), '')
    """
    s = str(val or '').strip()
    m = re.match(r'^(.*?\D)\s*([A-Za-z°%Ωµ]+)$', s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, ''
# --------------------------------------------------------------------------- #
# 5)  Streamlit UI                                                            #
# ---------------------------  
# Prompt เริ่มต้น เมื่อ user ไม่โหลดไฟล์ Excel มา
# … import ทั้งหมดเหมือนเดิม …

# Prompt เริ่มต้น (ต้องประกาศไว้ก่อน UI)
default_prompt = """กรุณาสกัดข้อมูลทั้งหมดจากรูปภาพนี้และแสดงผลในรูปแบบ JSON …"""

tab1, tab2 = st.tabs(["สกัดจากรูปภาพ","ประมวลผลจาก validated"])

# ------ โหมด 1: สกัดจากรูปภาพ ------
with tab1:
    st.subheader("💡 สกัดข้อมูลจากรูปภาพ")
    excel_f1 = st.file_uploader("1. เลือกไฟล์ Excel attributes", ["xlsx","xls"], key="tab1_attr")
    imgs    = st.file_uploader("2. อัปโหลดรูปภาพ (หลายไฟล์)", ["jpg","png","jpeg"],
                               accept_multiple_files=True, key="tab1_imgs")

    if st.button("ประมวลผลภาพ", key="btn_extract") and excel_f1 and imgs:
        # --- เอาโค้ดบล็อก extraction+wide มาใส่ที่นี่ ---
        prompt = default_prompt
        # ถ้ามี excel_f1 ให้ generate_prompt_from_excel…
        if excel_f1:
            prompt = generate_prompt_from_excel(excel_f1)
        st.expander("Prompt").write(prompt)

        # สกัดรูป → results …
        results, bar, status = [], st.progress(0), st.empty()
        for i,f in enumerate(imgs,1):
            bar.progress(i/len(imgs))
            status.write(f"กำลังประมวลผล {i}/{len(imgs)} – {f.name}")
            b64,mime = encode_image(f)
            js = {}
            resp = extract_data_from_image(API_KEY, b64, mime, prompt)
            try:
                js = json.loads(resp[resp.find('{'):resp.rfind('}')+1])
            except:
                js = {"error":resp}
            results.append({"file_name":f.name,"extracted_data":js})

        # เติม POWTR_CODE
        results = add_powtr_codes(results)

        # สร้าง df_long, pivot→df_wide, แสดง + ดาวน์โหลด wide
        rows = []
        for r in results:
            d = r["extracted_data"]
            assetnum = d.get("ASSETNUM","")
            siteid   = d.get("SITEID","")
            powtr    = d.get("POWTR_CODE","")
            if "error" in d or "raw_text" in d:
                rows.append({
                    "ASSETNUM":assetnum,
                    "SITEID":siteid,
                    "POWTR_CODE":powtr,
                    "ATTRIBUTE":"Error",
                    "VALUE":d.get("error",d.get("raw_text",""))
                })
            else:
                for attr,val in d.items():
                    if attr in ("ASSETNUM","SITEID","POWTR_CODE"): continue
                    rows.append({
                        "ASSETNUM":assetnum,
                        "SITEID":siteid,
                        "POWTR_CODE":powtr,
                        "ATTRIBUTE":attr,
                        "VALUE":val
                    })
        df_long = pd.DataFrame(rows)
        st.subheader("ตัวอย่างข้อมูล (แบบแถว)")
        st.dataframe(df_long)


        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df_long.to_excel(w, index=False)
        buf.seek(0)
        st.download_button("ดาวน์โหลด long", buf, "extracted_long.xlsx")

# ------ โหมด 2: ประมวลผลจาก validated ------
with tab2:
    st.subheader("🔍 ประมวลผลจากไฟล์ validated")
    validated_file = st.file_uploader(
        "1. เลือกไฟล์ validated_powtr_codes.xlsx", ["xlsx"], key="tab2_val"
    )
    template_file = st.file_uploader(
        "2. เลือกไฟล์ Template Excel", ["xlsx"], key="tab2_tmpl"
    )

    if st.button("ประมวลผล validated", key="btn_valid"):
        if not validated_file or not template_file:
            st.warning("กรุณาอัปโหลดทั้งไฟล์ validated และ Template Excel")
        else:
            # 1) อ่านและกรองเฉพาะแถวที่ Is_Correct==True
            df_val = pd.read_excel(validated_file)
            df_val = df_val[df_val["Is_Correct"] == True]
            st.subheader("Filtered validated (Is_Correct == True)")
            st.dataframe(df_val)

            # 2) สร้างตารางแบบ long เหมือน Tab1
            rows = []
            for _, row in df_val.iterrows():
                assetnum = row.get("ASSETNUM", "")
                siteid   = row.get("SITEID", "")
                powtr    = row.get("Correct_POWTR_CODE", "")

                # วนทุกคอลัมน์ที่ไม่ใช่ metadata
                for col in df_val.columns:
                    if col in [
                        "ASSETNUM", "SITEID",
                        "Is_Correct", "Correct_POWTR_CODE",
                        "Plant", "Location Description"
                    ]:
                        continue
                    raw = row[col]
                    # แยก value / unit
                    val, unit = split_value_unit(raw)
                    rows.append({
                        "ASSETNUM":     assetnum,
                        "SITEID":       siteid,
                        "POWTR_CODE":   powtr,
                        "ATTRIBUTE":    col,
                        "VALUE":        val,
                        "MEASUREUNIT":  unit
                    })

            df_long_val = pd.DataFrame(rows)
            st.subheader("ตัวอย่างข้อมูล (แบบแถว) จาก validated")
            st.dataframe(df_long_val)

            # 3) ให้ดาวน์โหลดตาราง long
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df_long_val.to_excel(writer, index=False)
            buf.seek(0)
            st.download_button(
                "ดาวน์โหลด extracted_long_from_validated.xlsx",
                buf,
                "extracted_long_from_validated.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_long_val"
            )

            # 4) (ถ้ายังต้องการ) เติมข้อมูลลงเทมเพลต
            try:
                out_buf = fill_template_from_validated(
                    validated_path=validated_file,
                    template_path=template_file
                )
                st.download_button(
                    "ดาวน์โหลด Template-POWTR-filled",
                    out_buf,
                    "Template-POWTR-filled.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_tmpl"
                )
            except Exception as e:
                st.error(f"ไม่สามารถสร้างไฟล์ตามเทมเพลตได้: {e}")