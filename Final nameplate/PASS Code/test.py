# --------------------  extract-excel.py  (FULL FILE – 27 Apr 2025)  --------------------
import os, base64, json, re, io, imghdr, requests, pandas as pd, streamlit as st
from PIL import Image
import io
import pandas as pd
from openpyxl import load_workbook

def fill_template_from_validated(
    validated_path: str,
    template_path: str,
    key_col_template: str = 'ASSETNUM',
    key_col_validated: str = 'Correct_POWTR_CODE',
    sheet_name: str = 'AssetAttr'
) -> io.BytesIO:
    """
    1) อ่าน validated_powtr_codes.xlsx → กรองเฉพาะ Is_Correct==True → index by Correct_POWTR_CODE  
    2) เปิดเทมเพลต → sheet AssetAttr → ลบแถวแรก (group header)  
    3) อ่าน header ใหม่ (row 1) → สร้าง map: header_name → col_index  
    4) สำหรับทุกแถว (row ≥ 2): lookup ASSETNUM ใน DataFrame → ถ้าพบ:
         • สำหรับทุกคอลัมน์ที่ชื่อซ้ำกับชื่อคอลัมน์ใน DF: เติมค่า rec[col]  
    5) เซฟ workbook ลง BytesIO แล้ว return
    """
    # 1) อ่าน & กรอง
    df = pd.read_excel(validated_path)
    df = df[df['Is_Correct'] == True].set_index(key_col_validated)

    # 2) โหลด template
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # 3) ลบแถว “group header”
    ws.delete_rows(1)

    # 4) อ่าน header ใหม่ เป็น dict(header_name→col_idx)
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    template_cols = {cell.value: cell.column for cell in header_row if cell.value}

    # ตรวจว่ามีคอลัมน์กุญแจ
    if key_col_template not in template_cols:
        raise KeyError(f"Template sheet ไม่มีคอลัมน์ '{key_col_template}'")

    # 5) เติมค่าในแต่ละแถว
    for row in ws.iter_rows(min_row=2):
        code = row[template_cols[key_col_template] - 1].value
        if code in df.index:
            rec = df.loc[code]
            # สำหรับทุก attribute ที่ DF มี และตรงกับ header ในเทมเพลต
            for attr, col_idx in template_cols.items():
                if attr in rec.index:
                    row[col_idx - 1].value = rec[attr]

    # 6) เซฟเป็น BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# --------------------------------------------------------------------------- #
# 1)  Utilities                                                               #
# --------------------------------------------------------------------------- #
from openpyxl import load_workbook

def fill_template_from_validated(validated_path: str,
                                 template_path: str) -> io.BytesIO:
    """
    • อ่าน validated_powtr_codes.xlsx ที่ Is_Correct==True  
    • โหลดเทมเพลต Copy of Template-…xlsx  
    • ลบแถวแรก (หัวข้อกลุ่ม)  
    • จับคู่ ASSETNUM กับ Correct_POWTR_CODE  
    • อัปเดต SITEID ← Plant, HIERARCHYPATH ← Location Description  
    • ส่งกลับไฟล์ Excel ในรูป BytesIO  
    """
    # อ่านและกรองเฉพาะแถวที่ติ๊กถูก
    df = pd.read_excel(validated_path)
    df = df[df['Is_Correct'] == True].set_index('Correct_POWTR_CODE')

    # โหลดเทมเพลต
    wb = load_workbook(template_path)
    ws = wb['AssetAttr']

    # ลบแถวแรก (แถวหัวข้อกลุ่ม)
    ws.delete_rows(1)

    # อ่าน header ใหม่ (แถว 1)
    header = next(ws.iter_rows(min_row=1, max_row=1))
    cols = {cell.value: cell.column for cell in header if cell.value}

    # ตรวจว่ามีคอลัมน์ที่ต้องใช้จริง
    for fld in ('ASSETNUM','SITEID','HIERARCHYPATH'):
        if fld not in cols:
            raise KeyError(f"Template sheet ไม่มีคอลัมน์ '{fld}'")

    # อัปเดตข้อมูลในแต่ละแถว (เริ่มที่ row 2)
    for row in ws.iter_rows(min_row=2):
        code = row[cols['ASSETNUM'] - 1].value
        if code in df.index:
            rec = df.loc[code]
            # แก้ ASSETNUM → Correct_POWTR_CODE
            row[cols['ASSETNUM']      -1].value = rec.name
            # SITEID ← Plant
            row[cols['SITEID']        -1].value = rec['Plant']
            # HIERARCHYPATH ← Location Description
            row[cols['HIERARCHYPATH'] -1].value = rec['Location Description']

    # เซฟลง BytesIO เพื่อส่งกลับ
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def encode_image(file) -> tuple[str, str]:
    """Convert an uploaded image file to (base64‑string, mime‑type) for Gemini."""
    raw = file.getvalue()
    kind = imghdr.what(None, raw) or 'jpeg'
    mime = f"image/{kind}"
    return base64.b64encode(raw).decode('utf-8'), mime


def _kv_from_text(txt: str) -> float | None:
    """
    Return the highest **system voltage** (kV) found in *txt*.

    • Accept “… 525000 V”, “34.5 kV”, “220 kV”, etc.  
    • Ignore values that are clearly power/current ratings (kVA, A, kA, VA).  
    • Skip numbers near “BIL” or “IMPULSE”.  
    • Discard absurdly large values > 1500 kV.
    """
    txt_u = txt.upper()
    best = None

    # split on '/', ',', ';' to isolate “33/30000/309000 kV” cases
    for chunk in re.split(r'[\/,;]', txt_u):
        chunk = chunk.strip()

        # skip chunks that mention KVA / VA / KA / A
        if re.search(r'\bK?VA\b|\bKA\b|\bAMP|\bA\b', chunk):
            continue

        # skip if near BIL / IMPULSE
        if 'BIL' in chunk or 'IMPULSE' in chunk:
            continue

        for m in re.finditer(r'(\d+(?:\.\d+)?)\s*([K]?V)(?![A-Z])', chunk):
            n = float(m.group(1))
            unit = m.group(2).upper()
            kv = n if unit == 'KV' else n / 1000
            if kv > 1500:          # absurdly high → ignore
                continue
            best = kv if best is None else max(best, kv)

    return best


# --------------------------------------------------------------------------- #
# 2)  Prompt generator                    #
# --------------------------------------------------------------------------- #
def generate_prompt_from_excel(excel_file):
    """
    Read an Excel list of attributes + (optionally) units, then build a Thai prompt
    telling Gemini to extract those exact fields in JSON.
    """
    # ----- read Excel whether it has a header row or not -----
    try:
        df = pd.read_excel(excel_file)
        first_col = df.columns[0]
        is_numeric_header = isinstance(first_col, (int, float))
        if is_numeric_header:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, header=None)
            df.columns = ['attribute_name'] + [f'col_{i}' for i in range(1, len(df.columns))]
            st.info("ตรวจพบไฟล์ไม่มีหัวคอลัมน์ – กำลังปรับให้อ่านได้")
    except Exception as e:
        excel_file.seek(0)
        df = pd.read_excel(excel_file, header=None)
        df.columns = ['attribute_name'] + [f'col_{i}' for i in range(1, len(df.columns))]
        st.warning(f"อ่านไฟล์แบบมีหัวคอลัมน์ไม่ได้: {e}  → ใช้โหมดไม่มีหัว")

    st.write("คอลัมน์ที่พบ:", list(df.columns))

    attribute_col = 'attribute_name'
    if attribute_col not in df.columns:
        for c in ['attribute_name', 'attribute', 'name', 'attributes',
                  'Attribute', 'ATTRIBUTE', 'field', 'Field', 'FIELD']:
            if c in df.columns:
                attribute_col = c; break
        if attribute_col not in df.columns:
            attribute_col = df.columns[0]
            st.warning(f"ไม่พบคอลัมน์ชื่อ attribute ที่รู้จัก – ใช้คอลัมน์ '{attribute_col}' แทน")

    unit_col = None
    for c in ['unit_of_measure', 'unit', 'Unit', 'UNIT', 'uom', 'UOM',
              'unit of measure', 'Unit of Measure']:
        if c in df.columns:
            unit_col = c; break

    if unit_col is None and len(df.columns) > 1:
        potential = df.columns[1]
        sample = df[potential].dropna().astype(str).tolist()[:10]
        if any(any(k in v for k in ['kg', 'V', 'A', 'kV', 'kVA', 'C', '°C',
                                    'mm', 'cm', 'm', '%']) for v in sample):
            unit_col = potential
            st.info(f"ตรวจพบคอลัมน์ '{potential}' อาจเป็นหน่วยวัด")

    prompt_parts = ["""กรุณาสกัดข้อมูลทั้งหมดจากรูปภาพนี้และแสดงผลในรูปแบบ JSON ที่มีโครงสร้างชัดเจน โดยใช้ key เป็นภาษาอังกฤษและ value เป็นข้อมูลที่พบ
ให้ return ค่า attributes กลับด้วยค่า attribute เท่านั้นห้าม return เป็น index เด็ดขาดและไม่ต้องเอาค่า index มาด้วย ให้ระวังเรื่อง voltage high side หน่วยต้องเป็น V หรือ kV เท่านั้น
โดยเอาเฉพาะ attributes ดังต่อไปนี้\n"""]

    for i, row in df.iterrows():
        attr = str(row[attribute_col]).strip()
        if pd.isna(attr) or attr == '':
            continue
        if unit_col and unit_col in df.columns and pd.notna(row[unit_col]) and str(row[unit_col]).strip():
            prompt_parts.append(f"{i+1}: {attr} [{row[unit_col]}]")
        else:
            prompt_parts.append(f"{i+1}: {attr}")

    prompt_parts.append("\nหากไม่พบข้อมูลสำหรับ attribute ใด ให้ใส่ค่า - แทน ไม่ต้องเดาค่า และให้รวม attribute และหน่วยวัดไว้ในค่าที่ส่งกลับด้วย")
    return "\n".join(prompt_parts)



# --------------------------------------------------------------------------- #
# 3)  Gemini API call                                                         #
# --------------------------------------------------------------------------- #
def extract_data_from_image(api_key: str, img_b64: str, mime: str, prompt: str) -> str:
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-04-17:generateContent?key={api_key}"
    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inlineData": {"mimeType": mime, "data": img_b64}}
            ]
        }],
        "generationConfig": {"temperature": 0.2, "topP": 0.85, "maxOutputTokens": 9000}
    }
    r = requests.post(url, headers={"Content-Type": "application/json"}, data=json.dumps(payload))
    if r.ok and r.json().get('candidates'):
        return r.json()['candidates'][0]['content']['parts'][0]['text']
    return f"API ERROR {r.status_code}: {r.text}"


# --------------------------------------------------------------------------- #
# 4)  POWTR‑CODE generator                                                    #
# --------------------------------------------------------------------------- #
def generate_powtr_code(extracted: dict) -> str:
    try:
# 1) Phase
        phase = '3'
        if any(any(k in str(v).upper() for k in ('1PH', '1-PH', 'SINGLE'))
               for v in extracted.values()):
            phase = '1'

        # 2) Voltage level
        high_kv = None
        for k, v in extracted.items():
            if any(t in k.upper() for t in ('VOLT', 'HV', 'LV', 'RATED', 'SYSTEM')):
                kv = _kv_from_text(str(v))
                if kv is not None:
                    high_kv = kv if high_kv is None else max(high_kv, kv)

        if high_kv is None:
            v_char = '-'
        elif high_kv > 765:
            return 'POWTR-3-OO'
        elif high_kv >= 345:
            v_char = 'E'
        elif high_kv >= 100:
            v_char = 'H'
        elif high_kv >= 1:
            v_char = 'M'
        else:
            v_char = 'L'

        # 3) Type → default = '-'  (เมื่อตรวจไม่เจอทั้ง DRY และ OIL)
        t_char = '-'
        for v in extracted.values():
            u = str(v).upper()
            if 'DRY' in u:
                t_char = 'D'
                break
            # ตรวจหา oil-based cooling class (OIL, ONAN, OFAF, ...)
            if any(kw in u for kw in ('OIL', 'ONAN', 'OFAF', 'OA', 'FOA')):
                t_char = 'O'
                break

        # 4) Tap‑changer (เหมือนเดิม)
        tap_char = 'F'
        for v in extracted.values():
            u = str(v).upper()
            if any(x in u for x in ('ON‑LOAD', 'ON-LOAD', 'OLTC')):
                tap_char = 'O'
                break
            if any(x in u for x in ('OFF‑LOAD', 'OFF-LOAD', 'FLTC', 'OCTC')):
                tap_char = 'F'

        return f'POWTR-{phase}{v_char}{t_char}{tap_char}'
    except Exception:
        return 'ไม่สามารถระบุได้'



def add_powtr_codes(results):
    for r in results:
        d = r.get('extracted_data', {})
        if isinstance(d, dict) and not any(k in d for k in ('error', 'raw_text')):
            d['POWTR_CODE'] = generate_powtr_code(d)
    return results


# --------------------------------------------------------------------------- #
# 5)  Streamlit UI                                                            #
# --------------------------------------------------------------------------- #
st.title("ระบบสกัดข้อมูลหม้อแปลง + POWTR‑CODE ")

API_KEY = "AIzaSyDb8iBV1EWqLvjheG_44gh3vQHfpmYGOCI"

tab1, tab2 = st.tabs(["ใช้ไฟล์ Excel", "ใช้ attributes ที่กำหนดไว้แล้ว"])
with tab1:
    excel_f = st.file_uploader("เลือกไฟล์ Excel attributes", ["xlsx", "xls"])
    if excel_f:
        st.dataframe(pd.read_excel(excel_f).head())
with tab2:
    use_def = st.checkbox("ใช้ attributes ที่กำหนดไว้แล้ว", True)
    if use_def:
        default_prompt = """กรุณาสกัดข้อมูลทั้งหมดจากรูปภาพนี้และแสดงผลในรูปแบบ JSON ที่มีโครงสร้างชัดเจน ..."""

imgs = st.file_uploader("อัปโหลดรูปภาพ (หลายไฟล์)", ["jpg", "png", "jpeg"],
                        accept_multiple_files=True)

if st.button("ประมวลผล") and API_KEY and imgs:
    prompt = default_prompt
    if 'excel_f' in locals() and excel_f:
        prompt = generate_prompt_from_excel(excel_f)
    st.expander("Prompt").write(prompt)

    results, bar, status = [], st.progress(0), st.empty()
    for i, f in enumerate(imgs, 1):
        bar.progress(i / len(imgs))
        status.write(f"กำลังประมวลผล {i}/{len(imgs)} – {f.name}")
        b64, mime = encode_image(f)
        resp = extract_data_from_image(API_KEY, b64, mime, prompt)

        try:
            js = json.loads(resp[resp.find('{'):resp.rfind('}') + 1])
        except Exception:
            js = {"error": resp}

        results.append({"file_name": f.name, "extracted_data": js})

    results = add_powtr_codes(results)
    # … after results = add_powtr_codes(results) …

    # 1) Build long‐form rows including ASSETNUM & SITEID
    rows = []
    for r in results:
        d = r['extracted_data']
        # pull out the two new key columns (or '' if missing)
        assetnum = d.get('ASSETNUM', '')
        siteid   = d.get('SITEID',   '')

        raw = d.get('POWTR_CODE', '')
        # prefix/backslash already applied in add_powtr_codes
        code = raw

        if 'error' in d or 'raw_text' in d:
            rows.append({
                'ASSETNUM': assetnum,
                'SITEID':   siteid,
                'POWTR_CODE': code,
                'ATTRIBUTE': 'Error',
                'VALUE': d.get('error', d.get('raw_text',''))
            })
        else:
            for k, v in d.items():
                if k == 'POWTR_CODE':
                    continue
                rows.append({
                    'ASSETNUM':   assetnum,
                    'SITEID':     siteid,
                    'POWTR_CODE': code,
                    'ATTRIBUTE':  k,
                    'VALUE':      v
                })

    df = pd.DataFrame(rows)
    st.subheader("ตัวอย่างข้อมูลที่สกัดได้ (แบบแถว)")
    st.dataframe(df)

    # 2) Pivot to one‐row‐per‐image, keeping ASSETNUM & SITEID at front
    df_wide = df.pivot(
        index=['ASSETNUM','SITEID','POWTR_CODE'],
        columns='ATTRIBUTE',
        values='VALUE'
    ).reset_index()
    # drop the columns name label
    df_wide.columns.name = None

    st.subheader("สรุปข้อมูลต่อรูปภาพ (wide format)")
    st.dataframe(df_wide)

    # 3) Download that wide table
    buf_wide = io.BytesIO()
    with pd.ExcelWriter(buf_wide, engine='openpyxl') as writer:
        df_wide.to_excel(writer, index=False)
    buf_wide.seek(0)
    st.download_button(
        "ดาวน์โหลดสรุปแบบ wide",
        buf_wide,
        "extracted_summary.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.subheader("POWTR‑CODE ที่สร้างได้")
    for r in results:
        st.write(r['extracted_data'].get('POWTR_CODE', '—'))

    rows = []
    for r in results:
        d = r['extracted_data']; code = d.get('POWTR_CODE', '')
        if 'error' in d or 'raw_text' in d:
            rows.append({"POWTR_CODE": code, "ATTRIBUTE": "Error",
                         "VALUE": d.get('error', d.get('raw_text', ''))})
        else:
            for k, v in d.items():
                if k != 'POWTR_CODE':
                    rows.append({"POWTR_CODE": code, "ATTRIBUTE": k, "VALUE": v})
    df = pd.DataFrame(rows)
    st.subheader("ตัวอย่างข้อมูลที่สกัดได้ (แบบแถว)")
    st.dataframe(df)

    # … (หลังสร้าง DataFrame df เรียบร้อย) …
    st.subheader("ดาวน์โหลดไฟล์ตามเทมเพลต")

template_path  = 'Copy of Template-MxLoader-Classification POW-TR.xlsx'
validated_path = 'validated_powtr_codes.xlsx'

try:
    out_buf = fill_template_from_validated(validated_path, template_path)
    st.download_button(
        label="ดาวน์โหลดไฟล์เติมข้อมูลตามเทมเพลต",
        data=out_buf,
        file_name="Template-POWTR-filled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
except Exception as e:
    st.error(f"ไม่สามารถสร้างไฟล์ตามเทมเพลตได้: {e}")
